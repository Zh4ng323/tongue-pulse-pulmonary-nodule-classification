# -*- coding: utf-8 -*-
"""
Pearson correlation analysis module.

Computes pairwise correlations with FDR correction and generates
heatmaps, clustermaps, and cross-modal correlation tables.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.colors as mcolors
import seaborn as sns
from scipy import stats
from scipy.cluster.hierarchy import linkage, dendrogram
from datetime import datetime
import os
import sys
import warnings

# 添加父目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

# 添加当前目录到路径，用于导入改进版函数
current_dir = os.path.dirname(os.path.abspath(__file__))
if current_dir not in sys.path:
    sys.path.insert(0, current_dir)

# 导入融合版绘图函数（统一优化版本）
from simple_correlation_improved_fusion import plot_corr_heatmap_improved, plot_clustermap_improved_v4

# 导入网络图模块
from correlation_network_plots import (
    plot_crossmodal_network,
    plot_differential_network,
    export_network_stats
)

warnings.filterwarnings('ignore')


# =============================================================================
# 自定义颜色映射
# =============================================================================

def create_coral_teal_colormap():
    """
    创建"橘红-白-青绿" (Coral-White-Teal) 配色方案（高饱和度版）

    Returns:
    --------
    LinearSegmentedColormap
        自定义颜色映射
    """
    # 定义颜色节点（提高饱和度）
    # -1.0: 橘红色 (Coral: #FF4500) - 加深
    # 0.0: 白色
    # 1.0: 青绿色 (Teal: #009688) - 加深
    colors = ['#FF4500', '#FFFFFF', '#009688']  # Coral → White → Teal
    nodes = [0.0, 0.5, 1.0]

    # 创建colormap
    cmap = mcolors.LinearSegmentedColormap.from_list('coral_teal', list(zip(nodes, colors)))

    return cmap


# 创建全局colormap实例
CORAL_TEAL_CMAP = create_coral_teal_colormap()


# =============================================================================
# 参数配置区
# =============================================================================

class SimpleCorrConfig:
    """简单相关性分析参数配置"""

    # 图形参数
    FIGURE_DPI = 300  # PNG 300dpi
    SAVE_FORMATS = ['png', 'svg']  # 移除PDF格式（仅保留PNG和SVG）
    FIGURE_SIZE_CORR = (16, 14)  # 舌×舌 / 脉×脉热图尺寸
    FIGURE_SIZE_CLUSTER = (14, 10)  # 聚类热图尺寸

    # 字体大小（增大以更清晰）
    FONT_SIZE_TITLE = 18
    FONT_SIZE_LABEL = 16
    FONT_SIZE_TICK = 13  # 加大
    FONT_SIZE_ANNOTATION = 12  # 加大

    # 热图配色：橘红-白-青绿 (Orange-White-Teal)
    CORMAP = 'RdBu_r'  # 临时使用，实际会创建自定义colormap
    VMIN, VMAX = -1, 1

    # 圆形气泡图参数
    CIRCLE_MIN_SIZE = 500  # 最小圆面积
    CIRCLE_MAX_SIZE = 5000  # 最大圆面积
    GRID_LINE_COLOR = '#CCCCCC'  # 浅灰色网格线
    GRID_LINE_WIDTH = 0.5  # 网格线宽度


# =============================================================================
# 工具函数
# =============================================================================

def compute_corr_p(df, apply_fdr=True):
    """
    计算Pearson相关系数矩阵和P值矩阵（含FDR校正）

    Parameters:
    -----------
    df : pd.DataFrame
        数据框（特征列）
    apply_fdr : bool, default=True
        是否应用FDR校正

    Returns:
    --------
    corr_matrix : pd.DataFrame
        相关系数矩阵 (r)
    p_matrix : pd.DataFrame
        P值矩阵（FDR校正后）
    p_raw_matrix : pd.DataFrame
        原始P值矩阵（用于表格导出）
    """
    cols = df.columns
    n = len(cols)

    corr_matrix = pd.DataFrame(np.eye(n), index=cols, columns=cols)
    p_matrix = pd.DataFrame(np.ones((n, n)), index=cols, columns=cols)
    p_raw_matrix = pd.DataFrame(np.ones((n, n)), index=cols, columns=cols)

    # 存储所有P值用于FDR校正
    all_p_values = []
    all_indices = []

    for i in range(n):
        for j in range(i+1, n):
            col1, col2 = cols[i], cols[j]

            # Pairwise删除缺失值：两列同时非空
            valid_idx = df[[col1, col2]].notna().all(axis=1)
            x = df.loc[valid_idx, col1]
            y = df.loc[valid_idx, col2]

            if len(x) < 3:
                # 样本量太小，无法计算
                r, p = np.nan, np.nan
            else:
                r, p = stats.pearsonr(x, y)

            corr_matrix.loc[col1, col2] = r
            corr_matrix.loc[col2, col1] = r
            p_raw_matrix.loc[col1, col2] = p
            p_raw_matrix.loc[col2, col1] = p

            if not np.isnan(p):
                all_p_values.append(p)
                all_indices.append((i, j))

    # 应用FDR校正
    if apply_fdr and len(all_p_values) > 0:
        from statsmodels.stats.multitest import multipletests
        rejected, p_corrected, _, _ = multipletests(
            all_p_values,
            alpha=0.05,
            method='fdr_bh'
        )

        for (p_corr, (i, j)) in zip(p_corrected, all_indices):
            col1, col2 = cols[i], cols[j]
            p_matrix.loc[col1, col2] = p_corr
            p_matrix.loc[col2, col1] = p_corr
    else:
        p_matrix = p_raw_matrix.copy()

    return corr_matrix, p_matrix, p_raw_matrix


def get_significance_stars_q(q):
    """
    根据q值（FDR校正后的P值）返回星号标记

    Parameters:
    -----------
    q : float
        q值（Benjamini-Hochberg FDR校正后的P值，即adjusted p-value）

    Returns:
    --------
    str
        星号标记 ('***', '**', '*', 或 '')
    """
    if pd.isna(q):
        return ''
    elif q < 0.001:
        return '***'
    elif q < 0.01:
        return '**'
    elif q < 0.05:
        return '*'
    else:
        return ''


def format_corr_with_stars(r, q):
    """
    格式化相关系数：r + 星号

    Parameters:
    -----------
    r : float
        相关系数
    q : float
        q值（FDR校正后的P值，即adjusted p-value）

    Returns:
    --------
    str
        格式化后的字符串，如 '0.15**'、'-0.30***'
    """
    if pd.isna(r):
        return 'N/A'

    stars = get_significance_stars_q(q)
    return f'{r:.2f}{stars}'


# =============================================================================
# 可视化函数
# =============================================================================

def plot_corr_bubble_heatmap(corr_matrix, q_matrix, title, output_path,
                             figsize=None, feature_type='tongue',
                             display_mode='value', show_grid=True):
    """
    绘制圆形气泡相关性热图（新版样式）

    样式要求：
    - 圆形气泡（大小+颜色双重编码相关强度）
    - 配色：橘红(负)→白(0)→青绿(正)
    - 可切换显示 r值 / q值星号
    - 添加浅灰色网格线

    Parameters:
    -----------
    corr_matrix : pd.DataFrame
        相关系数矩阵
    q_matrix : pd.DataFrame
        q值矩阵（FDR校正后的P值，即adjusted p-value）
    title : str
        图表标题
    output_path : str
        输出路径（不含扩展名）
    figsize : tuple, optional
        图形尺寸
    feature_type : str
        特征类型 ('tongue' 或 'pulse')
    display_mode : str, default='value'
        显示模式：'value' 显示r值，'stars' 显示显著性星号
    show_grid : bool, default=True
        是否显示网格线
    """
    if figsize is None:
        figsize = SimpleCorrConfig.FIGURE_SIZE_CORR

    n_vars = len(corr_matrix)
    fig, ax = plt.subplots(figsize=figsize)

    # 设置出版级样式
    plt.rcParams['font.size'] = 10
    plt.rcParams['font.family'] = config.FONT_ENGLISH
    plt.rcParams['axes.labelweight'] = 'bold'
    plt.rcParams['font.sans-serif'] = ['Arial', 'DejaVu Sans', 'Helvetica', 'sans-serif']

    # 准备数据
    corr_values = corr_matrix.values

    # 绘制圆形气泡
    for i in range(n_vars):
        for j in range(n_vars):
            if i == j:
                # 对角线：留空或显示特征名
                continue

            r = corr_values[i, j]

            if pd.isna(r):
                continue

            # 计算圆形半径（与|r|成正比，实现大小+颜色双重编码）
            abs_r = abs(r)
            rmin, rmax = 0.10, 0.45
            radius = rmin + abs_r * (rmax - rmin)

            # 绘制圆形
            circle = plt.Circle((j, i), radius=radius,  # 圆的半径（根据|r|变化）
                              color=CORAL_TEAL_CMAP((r + 1) / 2),  # 颜色映射到[-1, 1]
                              alpha=0.8)
            ax.add_patch(circle)

            # 添加文字标注
            q = q_matrix.iloc[i, j]

            if display_mode == 'value':
                # 显示r值
                text_color = 'white' if abs_r > 0.5 else 'black'
                ax.text(j, i, f'{r:.2f}',
                       ha='center', va='center',
                       color=text_color,
                       fontsize=SimpleCorrConfig.FONT_SIZE_ANNOTATION,
                       fontweight='normal')
            elif display_mode == 'stars':
                # 显示显著性星号
                stars = get_significance_stars_q(q)
                if stars:
                    ax.text(j, i, stars,
                           ha='center', va='center',
                           color='darkred',
                           fontsize=14,
                           fontweight='bold')
            else:  # 'both' 模式
                # 同时显示r值和星号
                text_color = 'white' if abs_r > 0.5 else 'black'
                stars = get_significance_stars_q(q)
                if stars:
                    ax.text(j, i, f'{r:.2f}\n{stars}',
                           ha='center', va='center',
                           color=text_color,
                           fontsize=SimpleCorrConfig.FONT_SIZE_ANNOTATION - 1,
                           fontweight='bold')
                else:
                    ax.text(j, i, f'{r:.2f}',
                           ha='center', va='center',
                           color=text_color,
                           fontsize=SimpleCorrConfig.FONT_SIZE_ANNOTATION,
                           fontweight='normal')

    # 添加网格线
    if show_grid:
        # 添加单元格网格线
        for i in range(n_vars + 1):
            ax.axhline(i - 0.5, color=SimpleCorrConfig.GRID_LINE_COLOR,
                      linewidth=SimpleCorrConfig.GRID_LINE_WIDTH, zorder=0)
            ax.axvline(i - 0.5, color=SimpleCorrConfig.GRID_LINE_COLOR,
                      linewidth=SimpleCorrConfig.GRID_LINE_WIDTH, zorder=0)

    # 设置坐标轴
    ax.set_xlim(-0.5, n_vars - 0.5)
    ax.set_ylim(n_vars - 0.5, -0.5)  # 反转y轴

    # 设置刻度和标签
    ax.set_xticks(range(n_vars))
    ax.set_yticks(range(n_vars))
    ax.set_xticklabels(corr_matrix.columns, fontsize=SimpleCorrConfig.FONT_SIZE_TICK,
                       rotation=45, ha='right')
    ax.set_yticklabels(corr_matrix.index, fontsize=SimpleCorrConfig.FONT_SIZE_TICK,
                       rotation=0)

    # 移除轴标题（保持简洁）
    ax.set_xlabel('')
    ax.set_ylabel('')

    # 添加colorbar
    sm = plt.cm.ScalarMappable(cmap=CORAL_TEAL_CMAP,
                               norm=plt.Normalize(vmin=-1, vmax=1))
    sm.set_array([])
    cbar = plt.colorbar(sm, ax=ax, fraction=0.046, pad=0.04)
    cbar.set_label('Pearson Correlation Coefficient (r)',
                   rotation=270, labelpad=20,
                   fontsize=SimpleCorrConfig.FONT_SIZE_LABEL, fontweight='bold')
    cbar.ax.tick_params(labelsize=SimpleCorrConfig.FONT_SIZE_TICK)

    # 设置标题
    ax.set_title(title, fontsize=SimpleCorrConfig.FONT_SIZE_TITLE,
                fontweight='bold', pad=15)

    # 设置白色背景
    ax.set_facecolor('white')
    fig.patch.set_facecolor('white')

    plt.tight_layout()

    # 保存图形
    for fmt in SimpleCorrConfig.SAVE_FORMATS:
        filepath = f'{output_path}.{fmt}'
        plt.savefig(filepath, dpi=SimpleCorrConfig.FIGURE_DPI,
                   bbox_inches='tight', facecolor='white')
        print(f"  [OK] 已保存: {os.path.basename(filepath)}")

    plt.close()


def plot_clustermap(corr_matrix, title, output_path, figsize=None,
                    display_values=True, dendrogram_color='#333333'):
    """
    绘制舌象×脉象聚类热图（图9 - 优化版）

    Parameters:
    -----------
    corr_matrix : pd.DataFrame
        相关系数矩阵（舌象×脉象）
    title : str
        图表标题
    output_path : str
        输出路径（不含扩展名）
    figsize : tuple, optional
        图形尺寸
    display_values : bool, default=True
        是否显示相关系数值
    dendrogram_color : str, default='#333333'
        树状图颜色（深灰色）
    """
    if figsize is None:
        figsize = SimpleCorrConfig.FIGURE_SIZE_CLUSTER

    # 使用seaborn clustermap
    g = sns.clustermap(
        corr_matrix,
        cmap=CORAL_TEAL_CMAP,  # 橘红-白-青绿
        vmin=SimpleCorrConfig.VMIN,
        vmax=SimpleCorrConfig.VMAX,
        figsize=figsize,
        row_cluster=True,
        col_cluster=True,
        method='average',  # 层次聚类方法
        metric='euclidean',
        linewidths=0.5,
        linecolor='white',
        cbar_kws={
            'label': 'Pearson Correlation Coefficient (r)',
            'orientation': 'vertical',
            'fraction': 0.046,
            'pad': 0.04
        }
    )

    # 设置标题
    g.fig.suptitle(title, fontsize=SimpleCorrConfig.FONT_SIZE_TITLE,
                   fontweight='bold', y=0.98)

    # 设置colorbar标签字体大小
    g.cax.tick_params(labelsize=SimpleCorrConfig.FONT_SIZE_TICK)

    # 调整刻度标签大小
    plt.setp(g.ax_heatmap.get_xticklabels(), fontsize=SimpleCorrConfig.FONT_SIZE_TICK,
             rotation=45, ha='right')
    plt.setp(g.ax_heatmap.get_yticklabels(), fontsize=SimpleCorrConfig.FONT_SIZE_TICK,
             rotation=0)

    # 设置树状图样式（细线条、深灰色）
    for dendro in [g.ax_row_dendrogram, g.ax_col_dendrogram]:
        if dendro is not None:
            # 设置树状图线条颜色和粗细
            for collection in dendro.collections:
                collection.set_color(dendrogram_color)
                collection.set_linewidth(1.0)  # 细线条

    # 添加相关系数标注（可选）
    if display_values:
        for i in range(len(corr_matrix.index)):
            for j in range(len(corr_matrix.columns)):
                r = corr_matrix.iloc[i, j]
                if not pd.isna(r):
                    text_color = 'white' if abs(r) > 0.5 else 'black'
                    g.ax_heatmap.text(j + 0.5, i + 0.5, f'{r:.2f}',
                                     ha='center', va='center',
                                     color=text_color,
                                     fontsize=SimpleCorrConfig.FONT_SIZE_ANNOTATION,
                                     fontweight='normal')

    # 移除轴标题（保持简洁）
    g.ax_heatmap.set_xlabel('')
    g.ax_heatmap.set_ylabel('')

    plt.tight_layout()

    # 保存图形
    for fmt in SimpleCorrConfig.SAVE_FORMATS:
        filepath = f'{output_path}.{fmt}'
        g.savefig(filepath, dpi=SimpleCorrConfig.FIGURE_DPI,
                bbox_inches='tight', facecolor='white')
        print(f"  [OK] 已保存: {os.path.basename(filepath)}")

    plt.close()


# =============================================================================
# 表格导出函数
# =============================================================================

def format_corr_table(corr_matrix, q_matrix):
    """
    格式化相关表格：r + 上标星号

    Parameters:
    -----------
    corr_matrix : pd.DataFrame
        相关系数矩阵
    q_matrix : pd.DataFrame
        q值矩阵（FDR校正后的P值，即adjusted p-value）

    Returns:
    --------
    pd.DataFrame
        格式化后的表格
    """
    formatted_df = corr_matrix.copy()

    for i in range(len(corr_matrix)):
        for j in range(len(corr_matrix.columns)):
            r = corr_matrix.iloc[i, j]
            q = q_matrix.iloc[i, j]

            if pd.isna(r):
                formatted_df.iloc[i, j] = 'N/A'
                continue

            stars = get_significance_stars_q(q)
            if stars:
                # 在文本格式中，星号直接跟在后面
                # Excel导出时会单独处理上标
                formatted_df.iloc[i, j] = f'{r:.2f}{stars}'
            else:
                formatted_df.iloc[i, j] = f'{r:.2f}'

    return formatted_df


def format_corr_table_excel(corr_matrix, q_matrix):
    """
    格式化相关表格：r + 上标星号（Excel富文本格式）

    Parameters:
    -----------
    corr_matrix : pd.DataFrame
        相关系数矩阵
    q_matrix : pd.DataFrame
        q值矩阵（FDR校正后的P值，即adjusted p-value）

    Returns:
    --------
    list of list
        格式化后的单元格数据（每个单元格是Cell对象）
    """
    from openpyxl.styles import Font

    # 创建数据表格（普通格式，包含星号）
    data = []
    for i in range(len(corr_matrix)):
        row = []
        for j in range(len(corr_matrix.columns)):
            r = corr_matrix.iloc[i, j]
            q = q_matrix.iloc[i, j]

            if pd.isna(r):
                row.append('N/A')
            else:
                stars = get_significance_stars_q(q)
                if stars:
                    # 文本格式：r + 星号（如 0.54***）
                    row.append(f'{r:.2f}{stars}')
                else:
                    row.append(f'{r:.2f}')
        data.append(row)

    return data


# =============================================================================
# 主分析类
# =============================================================================

class SimpleCorrelationAnalyzer:
    """
    简单相关性分析器

    功能：
    - 分组计算Pearson相关系数
    - 绘制舌×舌、脉×脉热图
    - 绘制舌×脉聚类热图
    - 导出相关表格
    """

    def __init__(self, data, target_col='Group'):
        """
        初始化分析器

        Parameters:
        -----------
        data : pd.DataFrame
            完整数据集
        target_col : str, default='Group'
            目标列名（用于分组）
        """
        self.data = data
        self.target_col = target_col

        # 分组数据
        self.groups = {
            'benign': data[data[target_col] == 0].copy(),  # 良性结节组
            'cancer': data[data[target_col] == 1].copy()   # 肺癌组
        }

        # 特征组
        self.feature_groups = {
            'tongue': [],
            'pulse': []
        }

        # 分析结果
        self.results = {
            'benign': {},
            'cancer': {}
        }

        print("\n" + "="*70)
        print("【模块1.5】简单相关性分析（Pearson相关）")
        print("="*70)
        print(f"\n[OK] 数据分组完成")
        print(f"  - 良性结节组: {len(self.groups['benign'])} 样本")
        print(f"  - 肺癌组: {len(self.groups['cancer'])} 样本")

    def auto_detect_features(self):
        """
        自动识别舌/脉特征（精确匹配）

        Returns:
        --------
        self : SimpleCorrelationAnalyzer
        """
        print("\n" + "-"*70)
        print("特征自动识别")
        print("-"*70)

        all_cols = self.data.columns.tolist()
        feature_cols = [col for col in all_cols if col != self.target_col]

        # 转换为小写进行匹配（处理大小写不一致）
        tongue_features_lower = [f.lower() for f in config.TONGUE_FEATURES]
        pulse_features_lower = [f.lower() for f in config.PULSE_FEATURES]

        # 识别舌象特征（精确匹配）
        tongue_features = []
        for col in feature_cols:
            if col.lower() in tongue_features_lower:
                tongue_features.append(col)

        # 识别脉象特征（精确匹配）
        pulse_features = []
        for col in feature_cols:
            if col.lower() in pulse_features_lower:
                pulse_features.append(col)

        self.feature_groups['tongue'] = tongue_features
        self.feature_groups['pulse'] = pulse_features

        print(f"\n[OK] 特征识别完成")
        print(f"  - 舌象特征: {len(tongue_features)}个")
        print(f"  - 脉象特征: {len(pulse_features)}个")

        return self

    def analyze_single_group(self, group_data, group_name):
        """
        对单组数据进行相关性分析

        FDR 校正策略：
        - 舌×舌：仅对舌象特征间的相关做 FDR 校正
        - 脉×脉：仅对脉象特征间的相关做 FDR 校正
        - 舌×脉：仅对舌象×脉象间的跨模态相关做 FDR 校正（cross-only 策略）

        Parameters:
        -----------
        group_data : pd.DataFrame
            单组数据
        group_name : str
            组名 ('benign' 或 'cancer')

        Returns:
        --------
        result : dict
            分析结果字典
        """
        print(f"\n{'-'*70}")
        print(f"正在分析: {group_name}")
        print(f"样本量: {len(group_data)}")
        print('-'*70)

        # 提取舌象和脉象特征
        tongue_features = self.feature_groups['tongue']
        pulse_features = self.feature_groups['pulse']

        # 检查特征是否存在于数据中
        tongue_features_exist = [f for f in tongue_features if f in group_data.columns]
        pulse_features_exist = [f for f in pulse_features if f in group_data.columns]

        skipped_tongue = set(tongue_features) - set(tongue_features_exist)
        skipped_pulse = set(pulse_features) - set(pulse_features_exist)

        if skipped_tongue:
            print(f"  [WARNING] 舌象特征被跳过（数据中不存在）: {skipped_tongue}")
        if skipped_pulse:
            print(f"  [WARNING] 脉象特征被跳过（数据中不存在）: {skipped_pulse}")

        print(f"  [INFO] 实际参与分析的舌象特征: {len(tongue_features_exist)}个")
        print(f"  [INFO] 实际参与分析的脉象特征: {len(pulse_features_exist)}个")

        # 计算舌象×舌象相关
        if len(tongue_features_exist) > 1:
            tongue_data = group_data[tongue_features_exist]
            corr_tt, p_tt, p_tt_raw = compute_corr_p(tongue_data, apply_fdr=True)
            print(f"  [OK] 舌象×舌象相关矩阵: {len(tongue_features_exist)}×{len(tongue_features_exist)}")
        else:
            corr_tt, p_tt, p_tt_raw = None, None, None
            print(f"  [SKIP] 舌象特征不足，跳过舌×舌分析")

        # 计算脉象×脉象相关
        if len(pulse_features_exist) > 1:
            pulse_data = group_data[pulse_features_exist]
            corr_pp, p_pp, p_pp_raw = compute_corr_p(pulse_data, apply_fdr=True)
            print(f"  [OK] 脉象×脉象相关矩阵: {len(pulse_features_exist)}×{len(pulse_features_exist)}")
        else:
            corr_pp, p_pp, p_pp_raw = None, None, None
            print(f"  [SKIP] 脉象特征不足，跳过脉×脉分析")

        # 计算舌象×脉象相关（cross-only FDR校正策略）
        # 仅对舌×脉跨模态相关做FDR校正，不包括舌×舌和脉×脉
        if len(tongue_features_exist) > 0 and len(pulse_features_exist) > 0:
            tongue_data = group_data[tongue_features_exist]
            pulse_data = group_data[pulse_features_exist]
            combined_data = pd.concat([tongue_data, pulse_data], axis=1)

            # Step 1: 计算原始 p 值（不做 FDR 校正）
            corr_all, _, p_raw_all = compute_corr_p(combined_data, apply_fdr=False)

            # Step 2: 提取舌×脉块的原始 p 值
            p_raw_tp = p_raw_all.loc[tongue_features_exist, pulse_features_exist].copy()

            # Step 3: 仅对舌×脉的 p 值做 BH-FDR 校正（cross-only 策略）
            from statsmodels.stats.multitest import multipletests
            p_values = p_raw_tp.values.flatten()
            p_values = p_values[~np.isnan(p_values)]  # 移除 NaN

            rejected, q_corrected, _, _ = multipletests(
                p_values,
                alpha=0.05,
                method='fdr_bh'
            )

            # Step 4: 回填 q 值到 DataFrame
            q_tp = p_raw_tp.copy()
            idx = 0
            for i in range(len(tongue_features_exist)):
                for j in range(len(pulse_features_exist)):
                    if not pd.isna(p_raw_tp.iloc[i, j]):
                        q_tp.iloc[i, j] = q_corrected[idx]
                        idx += 1

            # Step 5: 提取舌×脉相关系数矩阵
            corr_tp = corr_all.loc[tongue_features_exist, pulse_features_exist].copy()

            # 保存原始 p 值和 q 值
            p_tp_raw = p_raw_tp.copy()  # 原始 p 值（未校正）
            p_tp = q_tp  # FDR 校正后的 q 值（重用字段名，保持兼容性）

            print(f"  [OK] 舌象×脉象相关矩阵: {len(tongue_features_exist)}×{len(pulse_features_exist)}")
            print(f"  [INFO] FDR校正策略: cross-only（仅对舌×脉跨模态相关做校正）")
        else:
            corr_tp, p_tp, p_tp_raw = None, None, None
            print(f"  [SKIP] 特征不足，跳过舌×脉分析")

        result = {
            'corr_tt': corr_tt,
            'p_tt': p_tt,
            'p_tt_raw': p_tt_raw,
            'corr_pp': corr_pp,
            'p_pp': p_pp,
            'p_pp_raw': p_pp_raw,
            'corr_tp': corr_tp,
            'p_tp': p_tp,
            'p_tp_raw': p_tp_raw,
            'tongue_features': tongue_features_exist,
            'pulse_features': pulse_features_exist
        }

        return result

    def analyze_all_groups(self):
        """
        对所有组进行相关性分析

        Returns:
        --------
        self : SimpleCorrelationAnalyzer
        """
        # 先识别特征
        self.auto_detect_features()

        # 分析良性结节组
        if len(self.groups['benign']) > 0:
            self.results['benign'] = self.analyze_single_group(
                self.groups['benign'],
                '良性结节组'
            )

        # 分析肺癌组
        if len(self.groups['cancer']) > 0:
            self.results['cancer'] = self.analyze_single_group(
                self.groups['cancer'],
                '肺癌组'
            )

        return self

    def plot_all_figures(self, output_dir, display_mode='value'):
        """
        绘制所有图表

        Parameters:
        -----------
        output_dir : str
            输出目录（figures目录）
        display_mode : str, default='value'
            显示模式：
            - 'value': 显示相关系数r值
            - 'stars': 显示显著性星号
            - 'both': 同时显示r值和星号
        """
        print("\n" + "="*70)
        print("开始生成图表")
        print(f"显示模式: {display_mode}")
        print("="*70)

        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)

        # 计算tables目录（output_dir的兄弟目录）
        tables_dir = os.path.join(os.path.dirname(output_dir), 'tables')
        # 确保tables目录存在
        os.makedirs(tables_dir, exist_ok=True)

        for group_name in ['benign', 'cancer']:
            result = self.results.get(group_name)
            if result is None:
                continue

            group_label = "Benign" if group_name == 'benign' else "Cancer"
            group_dir = os.path.join(output_dir, group_name)
            os.makedirs(group_dir, exist_ok=True)

            # 确定A/B标签
            ab_label = '(A)' if group_name == 'benign' else '(B)'

            print(f"\n{'='*70}")
            print(f"Generating figures for {group_label} group")
            print('='*70)

            # Fig7: Tongue × Tongue correlation heatmap
            if result['corr_tt'] is not None:
                print("\n[Fig7] Tongue Features Correlation Heatmap")
                output_path = os.path.join(group_dir,
                                           f'Fig7_TongueCorr_{group_label}_{timestamp}')
                plot_corr_heatmap_improved(
                    corr_matrix=result['corr_tt'],
                    q_matrix=result['p_tt'],  # 实际是FDR校正后的q值
                    title=f'{ab_label} Tongue Features Correlation - {group_label}',
                    output_path=output_path,
                    feature_type='tongue'
                )

            # Fig8: Pulse × Pulse correlation heatmap
            if result['corr_pp'] is not None:
                print("\n[Fig8] Pulse Features Correlation Heatmap")
                output_path = os.path.join(group_dir,
                                           f'Fig8_PulseCorr_{group_label}_{timestamp}')
                plot_corr_heatmap_improved(
                    corr_matrix=result['corr_pp'],
                    q_matrix=result['p_pp'],  # 实际是FDR校正后的q值
                    title=f'{ab_label} Pulse Features Correlation - {group_label}',
                    output_path=output_path,
                    feature_type='pulse',
                    cell_scale=1.3  # 脉象格子大，圆点需要额外放大30%
                )

            # Fig9: Tongue × Pulse correlation clustermap
            if result['corr_tp'] is not None:
                print("\n[Fig9] Tongue x Pulse Correlation Clustermap")
                output_path = os.path.join(group_dir,
                                           f'Fig9_TonguePulseCluster_{group_label}_{timestamp}')
                plot_clustermap_improved_v4(  # ← 使用改进版v4函数（完全手动绘制+色条在右侧+树状图更粗+整体更扁）
                    corr_matrix=result['corr_tp'],
                    q_matrix=result['p_tp'],  # 实际是FDR校正后的q值
                    title=f'{ab_label} Tongue × Pulse Correlation Clustermap - {group_label}',
                    output_path=output_path
                )

                # ========== 新增: Fig9-NET 跨模态相关网络图 ==========
                print("\n[Fig9-NET] Tongue × Pulse Cross-modal Correlation Network")
                output_path_net = os.path.join(group_dir,
                                              f'Fig9Net_TonguePulseNetwork_{group_label}_{timestamp}')
                G, edges = plot_crossmodal_network(
                    corr_matrix=result['corr_tp'],
                    q_matrix=result['p_tp'],
                    tongue_features=result['tongue_features'],
                    pulse_features=result['pulse_features'],
                    group_name=group_label,
                    output_path=output_path_net
                )
                # 保存边数据到结果中（后续用于差异分析和统计表导出）
                result['network_edges'] = edges
                # =========================================

        print("\n" + "="*70)
        print("[OK] All figures have been saved")
        print("="*70)

        # ========== 新增: 生成差异网络图 (Fig9-DIFF) ==========
        print("\n" + "="*70)
        print("正在生成差异网络图 (Cancer vs Benign)...")
        print("="*70)

        # 确保两组都分析完成
        if (self.results.get('benign') and self.results.get('cancer') and
            self.results['benign'].get('corr_tp') is not None and
            self.results['cancer'].get('corr_tp') is not None):

            # 样本量
            n_benign = len(self.groups['benign'])
            n_cancer = len(self.groups['cancer'])

            # 输出路径 (diff图放在figures根目录)
            output_path_diff = os.path.join(output_dir,
                                          f'Fig9Diff_TonguePulseNetworkDiff_CancerVsBenign_{timestamp}')

            # 绘制差异网络图
            G_diff, diff_edges = plot_differential_network(
                benign_corr=self.results['benign']['corr_tp'],
                cancer_corr=self.results['cancer']['corr_tp'],
                benign_q=self.results['benign']['p_tp'],
                cancer_q=self.results['cancer']['p_tp'],
                tongue_features=self.results['benign']['tongue_features'],
                pulse_features=self.results['benign']['pulse_features'],
                n_benign=n_benign,
                n_cancer=n_cancer,
                output_path=output_path_diff
            )

            # 保存差异边数据
            self.results['diff_edges'] = diff_edges

            # 导出网络统计表
            stats_path = os.path.join(tables_dir, f'NetworkStats_{timestamp}')
            export_network_stats(
                benign_edges=self.results['benign'].get('network_edges', []),
                cancer_edges=self.results['cancer'].get('network_edges', []),
                diff_edges=diff_edges,
                n_benign=n_benign,
                n_cancer=n_cancer,
                output_path=stats_path
            )
        else:
            print("  [SKIP] 缺少必要的分析结果,跳过差异网络图生成")
        # =========================================

    def export_all_tables(self, output_dir):
        """
        导出所有表格

        Parameters:
        -----------
        output_dir : str
            输出目录
        """
        print("\n" + "="*70)
        print("开始导出表格")
        print("="*70)

        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)

        for group_name in ['benign', 'cancer']:
            result = self.results.get(group_name)
            if result is None:
                continue

            group_label = "Benign" if group_name == 'benign' else "Cancer"

            # Table 2 (Benign) and Table 3 (Cancer)
            if result['corr_tp'] is not None:
                table_num = 2 if group_name == 'benign' else 3
                print(f"\n[Table {table_num}] Tongue × Pulse Correlation - {group_label} Group")

                corr_matrix = result['corr_tp']
                q_matrix = result['p_tp']

                # 导出xlsx（带富文本格式）
                try:
                    from openpyxl import Workbook
                    from openpyxl.styles import Font, Alignment

                    xlsx_file = os.path.join(output_dir,
                                             f'Table{table_num}_TonguePulseCorr_{group_label}_{timestamp}.xlsx')

                    wb = Workbook()
                    ws = wb.active
                    ws.title = f'Table{table_num}_TonguePulseCorr'

                    # 写入行标题（舌象特征）
                    ws.cell(row=1, column=1).value = 'Tongue\\Pulse'
                    for j, col_name in enumerate(corr_matrix.columns, 2):
                        ws.cell(row=1, column=j).value = col_name
                        ws.cell(row=1, column=j).font = Font(bold=True)
                        ws.cell(row=1, column=j).alignment = Alignment(horizontal='center')

                    # 写入列标题（脉象特征）和数据
                    for i, row_name in enumerate(corr_matrix.index, 2):
                        ws.cell(row=i, column=1).value = row_name
                        ws.cell(row=i, column=1).font = Font(bold=True)
                        ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')

                        for j, col_name in enumerate(corr_matrix.columns, 2):
                            r = corr_matrix.loc[row_name, col_name]
                            q = q_matrix.loc[row_name, col_name]
                            p_raw = result['p_tp_raw'].loc[row_name, col_name]

                            if pd.isna(r):
                                ws.cell(row=i, column=j).value = 'N/A'
                            else:
                                stars = get_significance_stars_q(q)
                                if stars:
                                    # 写入富文本：r值 + 上标星号
                                    from openpyxl.cell.rich_text import CellRichText, TextBlock
                                    from openpyxl.cell.text import InlineFont
                                    from openpyxl.styles.colors import Color

                                    cell = ws.cell(row=i, column=j)

                                    # 创建 InlineFont 用于星号样式（上标、加粗、红色）
                                    inline_font = InlineFont(b=True, vertAlign='superscript')
                                    inline_font.color = Color(rgb="FFFF0000")

                                    # 创建 TextBlock：星号
                                    star_block = TextBlock(inline_font, stars)

                                    # 创建 CellRichText：纯字符串(r值) + TextBlock(星号)
                                    rich_text = CellRichText([f'{r:.2f}', star_block])

                                    cell.value = rich_text
                                else:
                                    ws.cell(row=i, column=j).value = f'{r:.2f}'

                            # 居中对齐
                            ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')

                    # 添加FDR校正说明
                    note_row = len(corr_matrix.index) + 3
                    ws.cell(row=note_row, column=1).value = 'Note:'
                    ws.cell(row=note_row, column=1).font = Font(italic=True, bold=True)
                    ws.cell(row=note_row + 1, column=1).value = '*** q<0.001, ** q<0.01, * q<0.05 (BH-FDR corrected)'
                    ws.cell(row=note_row + 1, column=1).font = Font(italic=True, size=10)
                    ws.cell(row=note_row + 2, column=1).value = 'FDR method: Benjamini-Hochberg (False Discovery Rate)'
                    ws.cell(row=note_row + 2, column=1).font = Font(italic=True, size=10)

                    # 调整列宽（修复超过26列崩溃的bug）
                    from openpyxl.utils import get_column_letter
                    ws.column_dimensions['A'].width = 20
                    for j in range(2, len(corr_matrix.columns) + 2):
                        ws.column_dimensions[get_column_letter(j)].width = 12

                    wb.save(xlsx_file)
                    print(f"  [OK] 已保存: {os.path.basename(xlsx_file)} (带上标星号)")

                except ImportError:
                    print("  [WARNING] openpyxl未安装，使用普通格式导出xlsx")
                    # 格式化表格
                    formatted_table = format_corr_table(corr_matrix, q_matrix)

                    xlsx_file = os.path.join(output_dir,
                                             f'Table{table_num}_TonguePulseCorr_{group_label}_{timestamp}.xlsx')
                    formatted_table.to_excel(xlsx_file, engine='openpyxl')
                    print(f"  [OK] 已保存: {os.path.basename(xlsx_file)}")

                # CSV导出已禁用（仅保留XLSX格式）
                # # 导出csv（纯文本格式）
                # formatted_table = format_corr_table(corr_matrix, p_matrix)
                # csv_file = os.path.join(output_dir,
                #                         f'Table{table_num}_TonguePulseCorr_{group_label}_{timestamp}.csv')
                # formatted_table.to_csv(csv_file, encoding='utf-8-sig')
                # print(f"  [OK] 已保存: {os.path.basename(csv_file)}")

            # 导出 Table S1-S4 (单模态相关：舌×舌、脉×脉)
            group_label = "Benign" if group_name == 'benign' else "Cancer"
            sample_size = len(self.groups[group_name])

            # Table S1/S3: 舌×舌相关
            if result['corr_tt'] is not None:
                table_num = 'S1' if group_name == 'benign' else 'S3'
                print(f"\n[Table {table_num}] Tongue × Tongue Self-Correlation - {group_label} Group")
                self.export_single_mode_correlation_table(
                    corr_matrix=result['corr_tt'],
                    q_matrix=result['p_tt'],
                    feature_names=result['tongue_features'],
                    table_number=table_num,
                    group_label=group_label,
                    output_dir=output_dir,
                    mode_name='tongue',
                    sample_size=sample_size
                )

            # Table S2/S4: 脉×脉相关
            if result['corr_pp'] is not None:
                table_num = 'S2' if group_name == 'benign' else 'S4'
                print(f"\n[Table {table_num}] Pulse × Pulse Self-Correlation - {group_label} Group")
                self.export_single_mode_correlation_table(
                    corr_matrix=result['corr_pp'],
                    q_matrix=result['p_pp'],
                    feature_names=result['pulse_features'],
                    table_number=table_num,
                    group_label=group_label,
                    output_dir=output_dir,
                    mode_name='pulse',
                    sample_size=sample_size
                )

        print("\n" + "="*70)
        print("[OK] All tables have been exported")
        print("="*70)

    def export_single_mode_correlation_table(
        self,
        corr_matrix,
        q_matrix,
        feature_names,
        table_number,
        group_label,
        output_dir,
        mode_name,
        sample_size
    ):
        """
        导出单模态相关表格（舌×舌、脉×脉）

        Parameters:
        -----------
        corr_matrix : pd.DataFrame
            相关系数矩阵 (n×n)
        q_matrix : pd.DataFrame
            q值矩阵（FDR校正后的P值，即adjusted p-value）
        feature_names : list
            特征名称列表
        table_number : str
            表格编号 ('S1', 'S2', 'S3', 'S4')
        group_label : str
            组别标签 ('Benign' or 'Cancer')
        output_dir : str
            输出目录
        mode_name : str
            模态名称 ('tongue' or 'pulse')
        sample_size : int
            样本量
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment

        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        mode_label = "Tongue" if mode_name == 'tongue' else "Pulse"

        xlsx_file = os.path.join(
            output_dir,
            f'Table{table_number}_{mode_label}SelfCorr_{group_label}_{timestamp}.xlsx'
        )

        wb = Workbook()
        ws = wb.active
        ws.title = f'Table{table_number}_{mode_label}SelfCorr'

        # 写入行标题和列标题
        for j, col_name in enumerate(feature_names, 2):
            ws.cell(row=1, column=j).value = col_name
            ws.cell(row=1, column=j).font = Font(bold=True)
            ws.cell(row=1, column=j).alignment = Alignment(horizontal='center')

        for i, row_name in enumerate(feature_names, 2):
            ws.cell(row=i, column=1).value = row_name
            ws.cell(row=i, column=1).font = Font(bold=True)
            ws.cell(row=i, column=1).alignment = Alignment(horizontal='center')

        # 写入数据（仅保留上三角，避免重复信息）
        for i, row_name in enumerate(feature_names, 2):
            for j, col_name in enumerate(feature_names, 2):
                # 跳过下三角（对称矩阵的下三角是重复信息）
                if j < i:
                    ws.cell(row=i, column=j).value = ''  # 留空
                    continue

                r = corr_matrix.loc[row_name, col_name]
                q = q_matrix.loc[row_name, col_name]

                if pd.isna(r):
                    ws.cell(row=i, column=j).value = 'N/A'
                else:
                    # 对角线显示1.00（无星号）
                    if i == j:
                        ws.cell(row=i, column=j).value = '1.00'
                    else:
                        stars = get_significance_stars_q(q)
                        # 使用富文本格式（上标星号）
                        if stars:
                            # 写入富文本：r值 + 上标星号
                            from openpyxl.cell.rich_text import CellRichText, TextBlock
                            from openpyxl.cell.text import InlineFont
                            from openpyxl.styles.colors import Color

                            cell = ws.cell(row=i, column=j)

                            # 创建 InlineFont 用于星号样式（上标、加粗、红色）
                            inline_font = InlineFont(b=True, vertAlign='superscript')
                            inline_font.color = Color(rgb="FFFF0000")

                            # 创建 TextBlock：星号
                            star_block = TextBlock(inline_font, stars)

                            # 创建 CellRichText：纯字符串(r值) + TextBlock(星号)
                            rich_text = CellRichText([f'{r:.2f}', star_block])

                            cell.value = rich_text
                        else:
                            ws.cell(row=i, column=j).value = f'{r:.2f}'

                # 居中对齐
                ws.cell(row=i, column=j).alignment = Alignment(horizontal='center')

        # 添加说明信息
        note_row = len(feature_names) + 3
        ws.cell(row=note_row, column=1).value = 'Note:'
        ws.cell(row=note_row, column=1).font = Font(italic=True, bold=True)
        ws.cell(row=note_row + 1, column=1).value = '*** q<0.001, ** q<0.01, * q<0.05 (BH-FDR corrected)'
        ws.cell(row=note_row + 1, column=1).font = Font(italic=True, size=10)
        ws.cell(row=note_row + 2, column=1).value = 'FDR method: Benjamini-Hochberg (False Discovery Rate)'
        ws.cell(row=note_row + 2, column=1).font = Font(italic=True, size=10)
        ws.cell(row=note_row + 3, column=1).value = f'Sample size: N={sample_size} ({group_label} group)'
        ws.cell(row=note_row + 3, column=1).font = Font(italic=True, size=10)

        # 调整列宽（修复超过26列崩溃的bug）
        from openpyxl.utils import get_column_letter
        ws.column_dimensions['A'].width = 20
        for j in range(2, len(feature_names) + 2):
            ws.column_dimensions[get_column_letter(j)].width = 12

        wb.save(xlsx_file)
        print(f"  [OK] 已保存: {os.path.basename(xlsx_file)}")

    def create_combined_figures(self, figures_dir):
        """
        创建benign和cancer的组合图片

        Parameters:
        -----------
        figures_dir : str
            图片目录（包含benign/和cancer/子目录）
        """
        from PIL import Image
        import glob

        # 定义要合并的图片类型 (新增Fig9Net)
        figure_types = ['Fig7', 'Fig8', 'Fig9', 'Fig9Net']

        for fig_type in figure_types:
            # 查找benign和cancer的图片（PNG格式）
            benign_pattern = os.path.join(figures_dir, 'benign', f'{fig_type}_*_Benign_*.png')
            cancer_pattern = os.path.join(figures_dir, 'cancer', f'{fig_type}_*_Cancer_*.png')

            benign_files = glob.glob(benign_pattern)
            cancer_files = glob.glob(cancer_pattern)

            if not benign_files or not cancer_files:
                print(f"  [SKIP] {fig_type}: 未找到配对的benign/cancer图片")
                continue

            # 取最新的文件
            benign_img_path = sorted(benign_files)[-1]
            cancer_img_path = sorted(cancer_files)[-1]

            # 打开图片
            img_benign = Image.open(benign_img_path)
            img_cancer = Image.open(cancer_img_path)

            # 获取图片尺寸
            width, height = img_benign.size

            # 创建新图片（宽度=2*原图宽度+间距，高度=原图高度）
            gap = 50  # 图片间距
            new_width = width * 2 + gap
            new_height = height
            combined_img = Image.new('RGB', (new_width, new_height), 'white')

            # 粘贴benign图片（左侧）
            combined_img.paste(img_benign, (0, 0))

            # 粘贴cancer图片（右侧）
            combined_img.paste(img_cancer, (width + gap, 0))

            # 保存合成图片（不添加A/B标签，标题中已包含）
            timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
            output_filename = f'{fig_type}_Combined_{timestamp}.png'
            output_path = os.path.join(figures_dir, output_filename)
            combined_img.save(output_path, dpi=(300, 300))
            print(f"  [OK] 已保存合成图: {output_filename}")

    def run_full_analysis(self, output_dir=None, display_mode='value'):
        """
        Run complete correlation analysis pipeline

        Parameters:
        -----------
        output_dir : str, optional
            Output directory
        display_mode : str, default='value'
            显示模式：'value', 'stars', 'both'
        """
        if output_dir is None:
            output_dir = os.path.join(config.OUTPUT_BASE_DIR, 'simple_correlation_results')

        # Create output directory
        os.makedirs(output_dir, exist_ok=True)

        print("\n" + "="*70)
        print("Starting Simple Correlation Analysis")
        print("="*70)

        # 分析所有组
        self.analyze_all_groups()

        # 创建输出目录（必须在调用plot_all_figures之前创建tables_dir）
        figures_dir = os.path.join(output_dir, 'figures')
        os.makedirs(figures_dir, exist_ok=True)

        tables_dir = os.path.join(output_dir, 'tables')
        os.makedirs(tables_dir, exist_ok=True)

        # 生成图表（plot_all_figures会导出网络统计到tables_dir）
        self.plot_all_figures(figures_dir, display_mode=display_mode)

        # 导出表格
        self.export_all_tables(tables_dir)

        # ========== 新增：生成合成图片 ==========
        print("\n" + "="*70)
        print("正在生成合成图片...")
        print("="*70)
        self.create_combined_figures(figures_dir)
        # ========================================

        print("\n" + "="*70)
        print("[OK] 简单相关性分析完成！")
        print("="*70)
        print(f"\n所有结果已保存到: {output_dir}")
        print(f"  - 图表: {figures_dir}")
        print(f"  - 表格: {tables_dir}")


# =============================================================================
# 便捷函数
# =============================================================================

def run_simple_correlation(data, target_col='Group', output_dir=None, display_mode='value'):
    """
    一键运行简单相关性分析

    Parameters:
    -----------
    data : pd.DataFrame
        完整数据集
    target_col : str, default='Group'
        目标列名
    output_dir : str, optional
        输出目录
    display_mode : str, default='value'
        显示模式：
        - 'value': 显示相关系数r值
        - 'stars': 显示显著性星号
        - 'both': 同时显示r值和星号

    Returns:
    --------
    analyzer : SimpleCorrelationAnalyzer
        分析器对象（包含所有结果）
    """
    analyzer = SimpleCorrelationAnalyzer(data, target_col)
    analyzer.run_full_analysis(output_dir, display_mode=display_mode)
    return analyzer


# =============================================================================
# 主程序测试
# =============================================================================

if __name__ == "__main__":
    print("\n" + "="*70)
    print("模块1.5测试：简单相关性分析")
    print("="*70)

    try:
        # 加载数据
        from modules.data_loader import DataLoader
        loader = DataLoader()
        loader.load_data()
        loader.auto_detect_features()

        # 运行相关性分析
        analyzer = run_simple_correlation(loader.data)

        print("\n[OK] 测试完成")

    except Exception as e:
        print(f"\n[X] 测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
