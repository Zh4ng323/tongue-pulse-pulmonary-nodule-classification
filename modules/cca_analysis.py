# -*- coding: utf-8 -*-
"""
Canonical Correlation Analysis (CCA) module.

Grouped CCA between tongue and pulse features with Bonferroni correction,
loading heatmaps, and scatter plots.
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from sklearn.cross_decomposition import CCA
from scipy import stats
from sklearn.preprocessing import StandardScaler
from datetime import datetime
import os
import sys
import warnings

# 添加父目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config

warnings.filterwarnings('ignore')


# =============================================================================
# 参数配置区
# =============================================================================

class CCAConfig:
    """CCA分析参数配置"""

    # 图形参数
    FIGURE_DPI = 600
    SAVE_FORMATS = ['jpg', 'svg']  # 移除PDF
    FIGURE_SIZE_CCA_SCATTER = (20, 8)
    FIGURE_SIZE_CCA_HEATMAP = (24, 12)  # 增大尺寸避免重叠

    # 字体大小（增大以更清晰）
    FONT_SIZE_TITLE = 18
    FONT_SIZE_LABEL = 16
    FONT_SIZE_TICK = 12
    FONT_SIZE_LEGEND = 12
    FONT_SIZE_HEATMAP_CELL = 12  # 热图格子里的数字字体

    # CCA分析参数
    N_COMPONENTS = None  # None表示自动确定
    SHOW_TOP_N = 3  # 显示前几对典型变量


# =============================================================================
# 代码实现
# =============================================================================


class GroupedCCAAnalyzer:
    """
    分组典型相关分析器

    功能：
    - 对两组数据（良性结节组、肺癌组）分别进行CCA
    - 生成符合期刊要求的图表
    - 导出详细的统计结果
    """

    def __init__(self, data, target_col='Group'):
        """
        初始化CCA分析器

        Parameters:
        -----------
        data : pd.DataFrame
            完整数据集
        target_col : str, default='Group'
            目标列名（用于分组）
        """
        self.data = data
        self.target_col = target_col

        # 分组数据
        self.groups = {
            'benign': data[data[target_col] == 0].copy(),  # 良性结节组
            'cancer': data[data[target_col] == 1].copy()   # 肺癌组
        }

        # 特征组
        self.feature_groups = {
            'tongue': [],
            'pulse': []
        }

        # 分析结果
        self.results = {
            'benign': {},
            'cancer': {}
        }

        print("\n" + "="*60)
        print("【模块2】典型相关分析（CCA）")
        print("="*60)
        print(f"\n[OK] 数据分组完成")
        print(f"  - 良性结节组: {len(self.groups['benign'])} 样本")
        print(f"  - 肺癌组: {len(self.groups['cancer'])} 样本")

    def auto_detect_features(self):
        """
        自动识别舌/脉特征（精确匹配）

        Returns:
        --------
        self : GroupedCCAAnalyzer
        """
        print("\n" + "-"*60)
        print("特征自动识别")
        print("-"*60)

        all_cols = self.data.columns.tolist()
        feature_cols = [col for col in all_cols if col != self.target_col]

        # 转换为小写进行匹配（处理大小写不一致）
        tongue_features_lower = [f.lower() for f in config.TONGUE_FEATURES]
        pulse_features_lower = [f.lower() for f in config.PULSE_FEATURES]

        # 识别舌象特征（精确匹配）
        tongue_features = []
        for col in feature_cols:
            if col.lower() in tongue_features_lower:
                tongue_features.append(col)

        # 识别脉象特征（精确匹配）
        pulse_features = []
        for col in feature_cols:
            if col.lower() in pulse_features_lower:
                pulse_features.append(col)

        self.feature_groups['tongue'] = tongue_features
        self.feature_groups['pulse'] = pulse_features

        print(f"\n[OK] 特征识别完成")
        print(f"  - 舌象特征 (X变量): {len(tongue_features)}个")
        print(f"  - 脉象特征 (Y变量): {len(pulse_features)}个")

        # 显示所有识别的特征
        if len(tongue_features) > 0:
            print(f"\n  舌象特征列表:")
            for i, feat in enumerate(tongue_features, 1):
                print(f"    {i}. {feat}")

        if len(pulse_features) > 0:
            print(f"\n  脉象特征列表:")
            for i, feat in enumerate(pulse_features, 1):
                print(f"    {i}. {feat}")

        return self

    def run_cca_single_group(self, group_data, group_name):
        """
        对单组数据进行CCA分析

        Parameters:
        -----------
        group_data : pd.DataFrame
            单组数据
        group_name : str
            组名 ('benign' 或 'cancer')

        Returns:
        --------
        result : dict
            CCA分析结果字典
        """
        print(f"\n{'-'*60}")
        print(f"正在分析: {group_name}")
        print('-'*60)

        # 提取舌象和脉象特征
        X = group_data[self.feature_groups['tongue']].values
        Y = group_data[self.feature_groups['pulse']].values

        # 检查数据有效性
        if X.shape[0] == 0 or Y.shape[0] == 0:
            print(f"[X] {group_name}: 没有有效数据")
            return None

        if X.shape[1] == 0 or Y.shape[1] == 0:
            print(f"[X] {group_name}: 特征数量为0")
            return None

        # 标准化
        scaler_X = StandardScaler()
        scaler_Y = StandardScaler()
        X_scaled = scaler_X.fit_transform(X)
        Y_scaled = scaler_Y.fit_transform(Y)

        # 确定CCA组数
        n_components = min(X.shape[1], Y.shape[1])

        # 执行CCA
        cca = CCA(n_components=n_components, max_iter=2000, tol=1e-6)
        cca.fit(X_scaled, Y_scaled)

        # 转换数据到典型空间
        X_c, Y_c = cca.transform(X_scaled, Y_scaled)

        # 计算典型相关系数
        corr_coefs = [np.corrcoef(X_c[:, i], Y_c[:, i])[0, 1] for i in range(n_components)]

        # 计算特征值
        eigenvalues = [(rho**2) / (1 - rho**2) for rho in corr_coefs]

        # 计算Wilks' Lambda和F统计量
        n = len(group_data)  # 样本量
        p = X.shape[1]       # X变量数
        q = Y.shape[1]       # Y变量数

        wilks_lambda = []
        f_statistics = []
        p_values = []

        for k in range(n_components):
            # Wilks' Lambda for kth component and beyond
            lambda_k = np.prod([1 - corr_coefs[j]**2 for j in range(k, n_components)])
            wilks_lambda.append(lambda_k)

            # F统计量（基于Rao的F近似）
            df1 = (p - k) * (q - k)
            df2 = n - 1 - (p + q + 1) / 2

            if df2 > 0:
                # Rao's F approximation
                s = np.sqrt((df1**2 * df2**2) / ((df1**2 + df2**2) - (p**2 + q**2 + 2) - 1))
                m = df1 * df2 - 2 / s
                t = (p * q - 2) / 4

                if k == 0:
                    f_stat = ((1 - lambda_k**(1/s)) / (lambda_k**(1/s))) * (m / t)
                else:
                    # 使用简化公式
                    f_stat = ((1 - lambda_k) / lambda_k) * ((n - 1 - (p + q + 1) / 2) / ((p - k) * (q - k)))
            else:
                f_stat = np.nan

            f_statistics.append(f_stat)

            # 计算P值（F检验）
            if not np.isnan(f_stat) and f_stat > 0:
                p_val = 1 - stats.f.cdf(f_stat, df1, max(df2, 1))
            else:
                p_val = np.nan

            p_values.append(p_val)

        # ========== Bonferroni校正 ==========
        p_values_raw = p_values.copy()
        n_tests = len(p_values)

        if n_tests > 0:
            # Bonferroni校正
            p_values_corrected = [min(p * n_tests, 1.0) if not np.isnan(p) else np.nan
                                  for p in p_values]
            # 替换原始P值
            p_values = p_values_corrected
        # =====================================

        # 计算结构系数（载荷）
        x_loadings = np.corrcoef(np.hstack((X_scaled, X_c)).T)[:X.shape[1], X.shape[1]:]
        y_loadings = np.corrcoef(np.hstack((Y_scaled, Y_c)).T)[:Y.shape[1], Y.shape[1]:]

        # 计算解释方差
        var_X = np.var(X_scaled, axis=0)
        var_Y = np.var(Y_scaled, axis=0)
        total_var_X = np.sum(var_X)
        total_var_Y = np.sum(var_Y)

        explained_var_X = np.var(X_c, axis=0) / total_var_X
        explained_var_Y = np.var(Y_c, axis=0) / total_var_Y

        cumulative_var_X = np.cumsum(explained_var_X)
        cumulative_var_Y = np.cumsum(explained_var_Y)

        # 打印结果
        print(f"\n[OK] CCA分析完成")
        print(f"  - 典型变量对数: {n_components}")
        print(f"\n  前三对典型变量:")
        for i in range(min(3, n_components)):
            sig = "***" if p_values[i] < 0.001 else "**" if p_values[i] < 0.01 else "*" if p_values[i] < 0.05 else ""
            print(f"    CV{i+1}: r={corr_coefs[i]:.3f}, 特征值={eigenvalues[i]:.3f}, F={f_statistics[i]:.2f}, "
                  f"p={p_values_raw[i]:.4f} (raw), p={p_values[i]:.4f} (Bonferroni-corrected) {sig}")

        # 保存结果
        result = {
            'n_components': n_components,
            'n_samples': len(group_data),
            'X_cols': self.feature_groups['tongue'],
            'Y_cols': self.feature_groups['pulse'],
            'corr_coefs': corr_coefs,
            'eigenvalues': eigenvalues,
            'f_statistics': f_statistics,
            'wilks_lambda': wilks_lambda,
            'p_values': p_values,
            'p_values_raw': p_values_raw,
            'n_tests': n_tests,
            'correction_method': 'Bonferroni',
            'x_loadings': x_loadings,
            'y_loadings': y_loadings,
            'X_c': X_c,
            'Y_c': Y_c,
            'explained_var_X': explained_var_X,
            'explained_var_Y': explained_var_Y,
            'cumulative_var_X': cumulative_var_X,
            'cumulative_var_Y': cumulative_var_Y,
            'cca_model': cca
        }

        return result

    def analyze_all_groups(self):
        """
        对所有组进行CCA分析

        Returns:
        --------
        self : GroupedCCAAnalyzer
        """
        # 先识别特征
        self.auto_detect_features()

        # 分析良性结节组
        if len(self.groups['benign']) > 0:
            self.results['benign'] = self.run_cca_single_group(
                self.groups['benign'],
                '良性结节组'
            )

        # 分析肺癌组
        if len(self.groups['cancer']) > 0:
            self.results['cancer'] = self.run_cca_single_group(
                self.groups['cancer'],
                '肺癌组'
            )

        return self

    def plot_scatter_bar(self, group_name, output_dir):
        """
        绘制散点图 + 相关系数条形图（图10/12风格）

        Parameters:
        -----------
        group_name : str
            组名 ('benign' 或 'cancer')
        output_dir : str
            输出目录
        """
        result = self.results.get(group_name)
        if result is None:
            return

        fig, (ax1, ax2) = plt.subplots(1, 2, figsize=CCAConfig.FIGURE_SIZE_CCA_SCATTER)

        # 设置出版级样式
        plt.rcParams['font.size'] = 11
        plt.rcParams['font.family'] = config.FONT_ENGLISH
        plt.rcParams['axes.labelweight'] = 'bold'
        plt.rcParams['axes.titlesize'] = 13

        # ========== 左图：典型变量散点图 ==========
        n_show = min(3, result['n_components'])
        colors = ['#1F77B4', '#FF7F0E', '#2CA02C']

        for i in range(n_show):
            # 格式化标签：左图显示 r 和 P值（无星号）
            rho = result["corr_coefs"][i]
            p_val = result["p_values"][i]

            # 格式化P值（校正后，无星号）
            if p_val < 0.001:
                p_str = "P<0.001"
            elif p_val > 0.999:
                p_str = "P>0.999"
            else:
                p_str = f"P={p_val:.3f}"

            ax1.scatter(result['X_c'][:, i], result['Y_c'][:, i],
                       alpha=0.6, color=colors[i],
                       label=f'CV{i+1} (r={rho:.2f}, {p_str})')

        ax1.set_xlabel('X Canonical Variables (Tongue)', fontweight='bold', fontsize=CCAConfig.FONT_SIZE_LABEL)
        ax1.set_ylabel('Y Canonical Variables (Pulse)', fontweight='bold', fontsize=CCAConfig.FONT_SIZE_LABEL)
        ax1.set_title('(A) Canonical Variables Scatter Plot', fontweight='bold', fontsize=CCAConfig.FONT_SIZE_TITLE)
        ax1.legend(loc='best', fontsize=CCAConfig.FONT_SIZE_LEGEND, title='Bonferroni-corrected')
        ax1.grid(True, alpha=0.3)

        # ========== 右图：相关系数条形图 ==========
        x_pos = np.arange(result['n_components'])
        bars = ax2.bar(x_pos, result['corr_coefs'], color='#1F77B4', alpha=0.8)
        ax2.axhline(y=config.CCA_R_LINE, color='red', linestyle='--', linewidth=2,
                   label=f'Threshold (r={config.CCA_R_LINE})', alpha=0.7)

        ax2.set_xticks(x_pos)
        ax2.set_xticklabels([f'CV{i+1}' for i in range(result['n_components'])], fontsize=CCAConfig.FONT_SIZE_TICK)
        ax2.set_ylabel('Canonical Correlation Coefficient', fontweight='bold', fontsize=CCAConfig.FONT_SIZE_LABEL)
        ax2.set_title('(B) Canonical Correlations Bar Plot', fontweight='bold', fontsize=CCAConfig.FONT_SIZE_TITLE)
        ax2.legend(loc='upper right', fontsize=CCAConfig.FONT_SIZE_LEGEND)
        ax2.grid(True, alpha=0.3, axis='y')

        # 添加数值标签：只显示R值和星号（基于校正后P值）
        for i, bar in enumerate(bars):
            height = bar.get_height()
            p_val = result['p_values'][i]  # Bonferroni校正后的P值

            # 显著性标记
            if p_val < 0.001:
                sig = "***"
            elif p_val < 0.01:
                sig = "**"
            elif p_val < 0.05:
                sig = "*"
            else:
                sig = ""

            ax2.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.2f}{sig}',
                    ha='center', va='bottom', fontsize=11, fontweight='bold')

        plt.tight_layout()

        # 保存图形（直接保存到group_dir，不创建子文件夹）
        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        group_label = "Benign" if group_name == 'benign' else "Cancer"
        filename = f'CCA_ScatterBar_{group_label}_{timestamp}'

        for fmt in CCAConfig.SAVE_FORMATS:
            filepath = os.path.join(output_dir, f'{filename}.{fmt}')
            plt.savefig(filepath, dpi=CCAConfig.FIGURE_DPI, bbox_inches='tight')
            print(f"  [OK] 已保存: {os.path.basename(filepath)}")

        plt.close()

    def plot_loadings_heatmap(self, group_name, output_dir):
        """
        绘制载荷热图（图11/13风格）- SCI级别优化版

        优化点：
        1. 显示所有数值（不筛选）
        2. 只显示前10对典型变量（和老师论文一致）
        3. 如果不足10对，则显示全部

        Parameters:
        -----------
        group_name : str
            组名 ('benign' 或 'cancer')
        output_dir : str
            输出目录
        """
        result = self.results.get(group_name)
        if result is None:
            return

        # 只显示前10对典型变量（与老师论文一致）
        n_display = min(10, result['n_components'])

        # 准备数据（只取前n_display列的载荷）
        x_loadings_df = pd.DataFrame(
            result['x_loadings'][:, :n_display],  # 只取前n_display列
            index=result['X_cols'],
            columns=[f'CV{i+1}' for i in range(n_display)]
        )

        y_loadings_df = pd.DataFrame(
            result['y_loadings'][:, :n_display],  # 只取前n_display列
            index=result['Y_cols'],
            columns=[f'CV{i+1}' for i in range(n_display)]
        )

        # 按第一典型变量的载荷绝对值排序
        x_loadings_df['abs_max'] = x_loadings_df.abs().max(axis=1)
        x_loadings_df = x_loadings_df.sort_values('abs_max', ascending=False)
        x_loadings_df = x_loadings_df.drop('abs_max', axis=1)

        y_loadings_df['abs_max'] = y_loadings_df.abs().max(axis=1)
        y_loadings_df = y_loadings_df.sort_values('abs_max', ascending=False)
        y_loadings_df = y_loadings_df.drop('abs_max', axis=1)

        # 创建图形 - 根据CV数量调整尺寸，增大避免重叠
        fig_height = max(12, n_display * 0.7)  # 增大高度
        fig = plt.figure(figsize=(26, fig_height))  # 增大宽度
        gs = fig.add_gridspec(1, 2, wspace=0.2,  # 增大子图间距
                              left=0.08, right=0.95, top=0.93, bottom=0.08)

        # ========== 左图：舌象特征载荷 ==========
        ax1 = fig.add_subplot(gs[0, 0])

        # 关闭默认网格
        ax1.grid(False)

        # 手动创建热图（不转置！）
        # 行=特征，列=CV
        im1 = ax1.imshow(x_loadings_df.values, cmap='RdBu_r', aspect='auto',
                        vmin=-1, vmax=1, interpolation='nearest')

        # 添加colorbar
        cbar1 = plt.colorbar(im1, ax=ax1, fraction=0.046, pad=0.04)
        cbar1.set_label('Loading Coefficient', rotation=270, labelpad=20,
                        fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        cbar1.ax.tick_params(labelsize=CCAConfig.FONT_SIZE_TICK)

        # 设置刻度和标签
        ax1.set_xticks(range(n_display))
        ax1.set_xticklabels(x_loadings_df.columns, fontsize=CCAConfig.FONT_SIZE_TICK, fontweight='bold', rotation=45, ha='right')
        ax1.set_yticks(range(len(x_loadings_df.index)))
        ax1.set_yticklabels(x_loadings_df.index, rotation=0, fontsize=CCAConfig.FONT_SIZE_TICK)

        # 轴标签和标题
        ax1.set_xlabel('Canonical Variables', fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        ax1.set_ylabel('Tongue Features', fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        ax1.set_title('(A) Tongue Features Loadings on Canonical Variables',
                     fontsize=CCAConfig.FONT_SIZE_TITLE, fontweight='bold', pad=20)

        # 添加所有数值标注（不筛选）
        for i in range(len(x_loadings_df.index)):  # 行（特征）
            for j in range(n_display):  # 列（CV）
                value = x_loadings_df.iloc[i, j]
                text_color = 'white' if abs(value) > 0.6 else 'black'
                weight = 'bold' if abs(value) > 0.5 else 'normal'
                ax1.text(j, i, f'{value:.2f}',
                        ha="center", va="center",
                        color=text_color, fontsize=CCAConfig.FONT_SIZE_HEATMAP_CELL, fontweight=weight)

        # ========== 右图：脉象特征载荷 ==========
        ax2 = fig.add_subplot(gs[0, 1])

        # 关闭默认网格
        ax2.grid(False)

        # 手动创建热图（不转置！）
        # 行=特征，列=CV
        im2 = ax2.imshow(y_loadings_df.values, cmap='RdBu_r', aspect='auto',
                        vmin=-1, vmax=1, interpolation='nearest')

        # 添加colorbar
        cbar2 = plt.colorbar(im2, ax=ax2, fraction=0.046, pad=0.04)
        cbar2.set_label('Loading Coefficient', rotation=270, labelpad=20,
                        fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        cbar2.ax.tick_params(labelsize=CCAConfig.FONT_SIZE_TICK)

        # 设置刻度和标签
        ax2.set_xticks(range(n_display))
        ax2.set_xticklabels(y_loadings_df.columns, fontsize=CCAConfig.FONT_SIZE_TICK, fontweight='bold', rotation=45, ha='right')
        ax2.set_yticks(range(len(y_loadings_df.index)))
        ax2.set_yticklabels(y_loadings_df.index, rotation=0, fontsize=CCAConfig.FONT_SIZE_TICK)

        # 轴标签和标题
        ax2.set_xlabel('Canonical Variables', fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        ax2.set_ylabel('Pulse Features', fontsize=CCAConfig.FONT_SIZE_LABEL, fontweight='bold')
        ax2.set_title('(B) Pulse Features Loadings on Canonical Variables',
                     fontsize=CCAConfig.FONT_SIZE_TITLE, fontweight='bold', pad=20)

        # 添加所有数值标注（不筛选）
        for i in range(len(y_loadings_df.index)):  # 行（特征）
            for j in range(n_display):  # 列（CV）
                value = y_loadings_df.iloc[i, j]
                text_color = 'white' if abs(value) > 0.6 else 'black'
                weight = 'bold' if abs(value) > 0.5 else 'normal'
                ax2.text(j, i, f'{value:.2f}',
                        ha="center", va="center",
                        color=text_color, fontsize=CCAConfig.FONT_SIZE_HEATMAP_CELL, fontweight=weight)

        plt.tight_layout()

        # 保存图形（直接保存到group_dir，不创建子文件夹）
        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        group_label = "Benign" if group_name == 'benign' else "Cancer"
        filename = f'CCA_LoadingsHeatmap_{group_label}_{timestamp}'

        for fmt in CCAConfig.SAVE_FORMATS:
            filepath = os.path.join(output_dir, f'{filename}.{fmt}')
            plt.savefig(filepath, dpi=CCAConfig.FIGURE_DPI, bbox_inches='tight',
                       facecolor='white', edgecolor='none')
            print(f"  [OK] 已保存: {os.path.basename(filepath)}")

        plt.close()

    def export_summary_table(self, group_name, output_dir):
        """
        导出统计表格（与原论文格式一致）

        Parameters:
        -----------
        group_name : str
            组名 ('benign' 或 'cancer')
        output_dir : str
            输出目录
        """
        result = self.results.get(group_name)
        if result is None:
            return

        # 格式化显著性标签（原始P值和校正后P值）
        significance_labels = []
        raw_p_labels = []

        for i, (p_raw, p_corr) in enumerate(zip(result['p_values_raw'], result['p_values'])):
            # 原始P值
            if p_raw < 0.001:
                raw_p_labels.append('P<0.001')
            elif p_raw > 0.999:
                raw_p_labels.append('P>0.999')
            else:
                raw_p_labels.append(f'{p_raw:.4f}')

            # 校正后P值（用于显著性判断）- 修复星号标注Bug
            if pd.isna(p_corr):
                significance_labels.append('N/A')
            elif p_corr < 0.001:
                significance_labels.append('P<0.001***')
            elif p_corr < 0.01:
                significance_labels.append(f'{p_corr:.4f}**')  # 修复：两颗星
            elif p_corr < 0.05:
                significance_labels.append(f'{p_corr:.4f}*')   # 修复：一颗星
            elif p_corr > 0.999:
                significance_labels.append('P>0.999')          # 修复：无星号
            else:
                significance_labels.append(f'{p_corr:.4f}')     # 修复：无星号（P ≥ 0.05）

        # 创建摘要表（包含原始和校正后P值）
        summary_data = {
            '序号': list(range(1, result['n_components'] + 1)),
            '相关性': [f'{r:.2f}' for r in result['corr_coefs']],
            '特征值': [f'{e:.2f}' for e in result['eigenvalues']],
            'F': [f'{f:.2f}' if not np.isnan(f) else 'N/A' for f in result['f_statistics']],
            '原始P值': raw_p_labels,
            'Bonferroni校正P值': significance_labels
        }

        summary_df = pd.DataFrame(summary_data)

        # 保存Excel（直接保存到group_dir，不创建子文件夹）
        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        group_label = "Benign" if group_name == 'benign' else "Cancer"
        excel_file = os.path.join(output_dir, f'CCA_SummaryTable_{group_label}_{timestamp}.xlsx')

        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            from openpyxl.styles import Font, Alignment

            # 摘要表（主表，与原论文格式一致）
            summary_df.to_excel(writer, sheet_name='Summary', index=False)

            # 在摘要表底部添加说明
            summary_ws = writer.sheets['Summary']
            note_row = len(summary_df) + 3
            summary_ws.cell(row=note_row, column=1).value = 'Note:'
            summary_ws.cell(row=note_row, column=1).font = Font(italic=True, bold=True)
            summary_ws.cell(row=note_row + 1, column=1).value = '*** P-values adjusted using Bonferroni correction'
            summary_ws.cell(row=note_row + 1, column=1).font = Font(italic=True, size=10)
            summary_ws.cell(row=note_row + 2, column=1).value = f'Number of tests: {result["n_tests"]}'
            summary_ws.cell(row=note_row + 2, column=1).font = Font(italic=True, size=10)

            # 详细统计表（包含Wilks' Lambda等）
            detail_data = {
                'Canonical Pair': [f'CV{i+1}' for i in range(result['n_components'])],
                'Correlation': result['corr_coefs'],
                'Eigenvalue': result['eigenvalues'],
                'F_Statistic': result['f_statistics'],
                'Wilks_Lambda': result['wilks_lambda'],
                'P_Value_Raw': result['p_values_raw'],
                'P_Value_Corrected': result['p_values'],
                'Significant': [p < 0.05 for p in result['p_values']]
            }
            detail_df = pd.DataFrame(detail_data)
            detail_df.to_excel(writer, sheet_name='Detailed_Stats', index=False)

            # X载荷
            x_loadings_df = pd.DataFrame(
                result['x_loadings'],
                index=result['X_cols'],
                columns=[f'X_CV{i+1}' for i in range(result['n_components'])]
            )
            x_loadings_df.to_excel(writer, sheet_name='X_Loadings')

            # Y载荷
            y_loadings_df = pd.DataFrame(
                result['y_loadings'],
                index=result['Y_cols'],
                columns=[f'Y_CV{i+1}' for i in range(result['n_components'])]
            )
            y_loadings_df.to_excel(writer, sheet_name='Y_Loadings')

            # 典型变量得分
            scores_df = pd.concat([
                pd.DataFrame(result['X_c'], columns=[f'X_CV{i+1}' for i in range(result['n_components'])]),
                pd.DataFrame(result['Y_c'], columns=[f'Y_CV{i+1}' for i in range(result['n_components'])]),
            ], axis=1)
            scores_df.to_excel(writer, sheet_name='Scores', index=False)

        print(f"  [OK] 已保存: {os.path.basename(excel_file)}")

    def export_loadings_tables(self, group_name, output_dir):
        """
        单独导出loadings数值表（用于论文补充材料）

        Parameters:
        -----------
        group_name : str
            组名 ('benign' 或 'cancer')
        output_dir : str
            输出目录
        """
        result = self.results.get(group_name)
        if result is None:
            return

        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        group_label = "Benign" if group_name == 'benign' else "Cancer"

        # ========== 导出X载荷（舌象特征）==========
        x_loadings_df = pd.DataFrame(
            result['x_loadings'],
            index=result['X_cols'],
            columns=[f'CV{i+1}' for i in range(result['n_components'])]
        )

        x_loadings_file = os.path.join(
            output_dir,
            f'Table_SX1_Loadings_Tongue_{group_label}_{timestamp}.xlsx'
        )

        # 使用openpyxl设置格式
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(x_loadings_file, engine='openpyxl') as writer:
            x_loadings_df.to_excel(writer, sheet_name='Tongue_Loadings', index=True)

            # 获取工作表并设置格式
            ws = writer.sheets['Tongue_Loadings']

            # 设置列宽
            ws.column_dimensions['A'].width = 25
            for j in range(2, len(x_loadings_df.columns) + 2):
                ws.column_dimensions[get_column_letter(j)].width = 14

            # 设置标题行样式
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

            # 设置行标题样式（特征名称）
            for i in range(2, len(x_loadings_df) + 2):
                ws.cell(row=i, column=1).font = Font(bold=True, size=11)
                ws.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')

            # 设置数值单元格格式
            for i in range(2, len(x_loadings_df) + 2):
                for j in range(2, len(x_loadings_df.columns) + 2):
                    cell = ws.cell(row=i, column=j)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 根据数值设置颜色（正数为绿色，负数为红色）
                    value = cell.value
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color='008000', size=10)  # 绿色
                        elif value < 0:
                            cell.font = Font(color='CC0000', size=10)  # 红色

            # 添加说明信息
            note_row = len(x_loadings_df) + 3
            ws.cell(row=note_row, column=1).value = 'Table S.X1: Tongue Features Loadings on Canonical Variables'
            ws.cell(row=note_row, column=1).font = Font(bold=True, size=12, italic=True)
            ws.cell(row=note_row + 1, column=1).value = f'Group: {group_label}'
            ws.cell(row=note_row + 1, column=1).font = Font(italic=True, size=11)
            ws.cell(row=note_row + 2, column=1).value = f'Sample size: N={result["n_samples"]}'
            ws.cell(row=note_row + 2, column=1).font = Font(italic=True, size=10)
            ws.cell(row=note_row + 3, column=1).value = 'Note: Loadings represent structure correlations between original variables and canonical variables'
            ws.cell(row=note_row + 3, column=1).font = Font(italic=True, size=10)

        print(f"  [OK] 已保存舌象载荷表: {os.path.basename(x_loadings_file)}")

        # ========== 导出Y载荷（脉象特征）==========
        y_loadings_df = pd.DataFrame(
            result['y_loadings'],
            index=result['Y_cols'],
            columns=[f'CV{i+1}' for i in range(result['n_components'])]
        )

        y_loadings_file = os.path.join(
            output_dir,
            f'Table_SX2_Loadings_Pulse_{group_label}_{timestamp}.xlsx'
        )

        with pd.ExcelWriter(y_loadings_file, engine='openpyxl') as writer:
            y_loadings_df.to_excel(writer, sheet_name='Pulse_Loadings', index=True)

            # 获取工作表并设置格式
            ws = writer.sheets['Pulse_Loadings']

            # 设置列宽
            ws.column_dimensions['A'].width = 25
            for j in range(2, len(y_loadings_df.columns) + 2):
                ws.column_dimensions[get_column_letter(j)].width = 14

            # 设置标题行样式
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

            # 设置行标题样式（特征名称）
            for i in range(2, len(y_loadings_df) + 2):
                ws.cell(row=i, column=1).font = Font(bold=True, size=11)
                ws.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')

            # 设置数值单元格格式
            for i in range(2, len(y_loadings_df) + 2):
                for j in range(2, len(y_loadings_df.columns) + 2):
                    cell = ws.cell(row=i, column=j)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 根据数值设置颜色（正数为绿色，负数为红色）
                    value = cell.value
                    if isinstance(value, (int, float)):
                        if value > 0:
                            cell.font = Font(color='008000', size=10)  # 绿色
                        elif value < 0:
                            cell.font = Font(color='CC0000', size=10)  # 红色

            # 添加说明信息
            note_row = len(y_loadings_df) + 3
            ws.cell(row=note_row, column=1).value = 'Table S.X2: Pulse Features Loadings on Canonical Variables'
            ws.cell(row=note_row, column=1).font = Font(bold=True, size=12, italic=True)
            ws.cell(row=note_row + 1, column=1).value = f'Group: {group_label}'
            ws.cell(row=note_row + 1, column=1).font = Font(italic=True, size=11)
            ws.cell(row=note_row + 2, column=1).value = f'Sample size: N={result["n_samples"]}'
            ws.cell(row=note_row + 2, column=1).font = Font(italic=True, size=10)
            ws.cell(row=note_row + 3, column=1).value = 'Note: Loadings represent structure correlations between original variables and canonical variables'
            ws.cell(row=note_row + 3, column=1).font = Font(italic=True, size=10)

        print(f"  [OK] 已保存脉象载荷表: {os.path.basename(y_loadings_file)}")

        # ========== 导出合并的loadings表（舌象+脉象）==========
        combined_file = os.path.join(
            output_dir,
            f'Table_SX3_Loadings_Combined_{group_label}_{timestamp}.xlsx'
        )

        with pd.ExcelWriter(combined_file, engine='openpyxl') as writer:
            # 舌象载荷在第一个工作表
            x_loadings_df.to_excel(writer, sheet_name='Tongue_Loadings', index=True)
            ws1 = writer.sheets['Tongue_Loadings']
            for col in range(1, len(x_loadings_df.columns) + 2):
                ws1.column_dimensions[get_column_letter(col)].width = 14

            # 脉象载荷在第二个工作表
            y_loadings_df.to_excel(writer, sheet_name='Pulse_Loadings', index=True)
            ws2 = writer.sheets['Pulse_Loadings']
            for col in range(1, len(y_loadings_df.columns) + 2):
                ws2.column_dimensions[get_column_letter(col)].width = 14

            # 添加说明工作表
            info_df = pd.DataFrame({
                'Description': [
                    'Canonical Correlation Analysis (CCA) - Loadings Tables',
                    f'',
                    f'Group: {group_label}',
                    f'Sample size: N={result["n_samples"]}',
                    f'Number of canonical components: {result["n_components"]}',
                    f'',
                    f'Tongue features (X variables): {len(result["X_cols"])}',
                    f'Pulse features (Y variables): {len(result["Y_cols"])}',
                    f'',
                    f'Note: Loadings (structure correlations) represent correlations between',
                    f'original variables and their canonical variables. Values range from -1 to 1.',
                    f'Higher absolute values indicate stronger contributions to the canonical relationship.'
                ]
            })
            info_df.to_excel(writer, sheet_name='README', index=False, header=False)

        print(f"  [OK] 已保存合并载荷表: {os.path.basename(combined_file)}")

    def plot_combined_comparison(self, output_dir):
        """
        绘制CCA对照四联图 (Benign vs Cancer)

        2x2布局:
        - (A) Benign: CV1-3散点图
        - (B) Cancer: CV1-3散点图
        - (C) Benign: Canonical correlations条形图
        - (D) Cancer: Canonical correlations条形图

        Parameters:
        -----------
        output_dir : str
            输出目录
        """
        print("\n[CCA对照四联图] 正在生成Benign vs Cancer对照图...")

        # 检查是否有两组结果
        if not self.results.get('benign') or not self.results.get('cancer'):
            print("  [SKIP] 缺少必要的分析结果")
            return

        result_benign = self.results['benign']
        result_cancer = self.results['cancer']

        # 创建2x2图形
        fig = plt.figure(figsize=(24, 10))
        gs = fig.add_gridspec(2, 2, hspace=0.3, wspace=0.3,
                              left=0.08, right=0.95, top=0.93, bottom=0.08)

        # ========== (A) Benign: CV1-3散点图 ==========
        ax1 = fig.add_subplot(gs[0, 0])
        n_show = min(3, result_benign['n_components'])
        colors = ['#1F77B4', '#FF7F0E', '#2CA02C']

        for i in range(n_show):
            rho = result_benign["corr_coefs"][i]
            p_val = result_benign["p_values"][i]

            if p_val < 0.001:
                p_str = "P<0.001"
            elif p_val > 0.999:
                p_str = "P>0.999"
            else:
                p_str = f"P={p_val:.3f}"

            ax1.scatter(result_benign['X_c'][:, i], result_benign['Y_c'][:, i],
                       alpha=0.6, color=colors[i],
                       label=f'CV{i+1} (r={rho:.2f}, {p_str})')

        ax1.set_xlabel('X Canonical Variables (Tongue)', fontweight='bold', fontsize=14)
        ax1.set_ylabel('Y Canonical Variables (Pulse)', fontweight='bold', fontsize=14)
        ax1.set_title('(A) Benign: Canonical Variables Scatter Plot', fontweight='bold', fontsize=16)
        ax1.legend(loc='best', fontsize=11)
        ax1.grid(True, alpha=0.3)

        # ========== (B) Cancer: CV1-3散点图 ==========
        ax2 = fig.add_subplot(gs[0, 1])
        for i in range(n_show):
            rho = result_cancer["corr_coefs"][i]
            p_val = result_cancer["p_values"][i]

            if p_val < 0.001:
                p_str = "P<0.001"
            elif p_val > 0.999:
                p_str = "P>0.999"
            else:
                p_str = f"P={p_val:.3f}"

            ax2.scatter(result_cancer['X_c'][:, i], result_cancer['Y_c'][:, i],
                       alpha=0.6, color=colors[i],
                       label=f'CV{i+1} (r={rho:.2f}, {p_str})')

        ax2.set_xlabel('X Canonical Variables (Tongue)', fontweight='bold', fontsize=14)
        ax2.set_ylabel('Y Canonical Variables (Pulse)', fontweight='bold', fontsize=14)
        ax2.set_title('(B) Cancer: Canonical Variables Scatter Plot', fontweight='bold', fontsize=16)
        ax2.legend(loc='best', fontsize=11)
        ax2.grid(True, alpha=0.3)

        # ========== (C) Benign: Canonical correlations条形图 ==========
        ax3 = fig.add_subplot(gs[1, 0])
        x_pos = np.arange(result_benign['n_components'])
        bars_benign = ax3.bar(x_pos, result_benign['corr_coefs'], color='#1F77B4', alpha=0.8)
        ax3.axhline(y=config.CCA_R_LINE, color='red', linestyle='--', linewidth=2,
                   label=f'Threshold (r={config.CCA_R_LINE})', alpha=0.7)

        ax3.set_xticks(x_pos)
        ax3.set_xticklabels([f'CV{i+1}' for i in range(result_benign['n_components'])], fontsize=12)
        ax3.set_ylabel('Canonical Correlation Coefficient', fontweight='bold', fontsize=14)
        ax3.set_title('(C) Benign: Canonical Correlations', fontweight='bold', fontsize=16)
        ax3.legend(loc='upper right', fontsize=11)
        ax3.grid(True, alpha=0.3, axis='y')
        ax3.set_ylim(0, 1.0)  # 固定y轴范围便于比较

        # 添加数值标签和星号
        for i, bar in enumerate(bars_benign):
            height = bar.get_height()
            p_val = result_benign['p_values'][i]

            if p_val < 0.001:
                sig = "***"
            elif p_val < 0.01:
                sig = "**"
            elif p_val < 0.05:
                sig = "*"
            else:
                sig = ""

            ax3.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.2f}{sig}',
                    ha='center', va='bottom', fontsize=11, fontweight='bold')

        # ========== (D) Cancer: Canonical correlations条形图 ==========
        ax4 = fig.add_subplot(gs[1, 1])
        x_pos = np.arange(result_cancer['n_components'])
        bars_cancer = ax4.bar(x_pos, result_cancer['corr_coefs'], color='#FF7F0E', alpha=0.8)
        ax4.axhline(y=config.CCA_R_LINE, color='red', linestyle='--', linewidth=2,
                   label=f'Threshold (r={config.CCA_R_LINE})', alpha=0.7)

        ax4.set_xticks(x_pos)
        ax4.set_xticklabels([f'CV{i+1}' for i in range(result_cancer['n_components'])], fontsize=12)
        ax4.set_ylabel('Canonical Correlation Coefficient', fontweight='bold', fontsize=14)
        ax4.set_title('(D) Cancer: Canonical Correlations', fontweight='bold', fontsize=16)
        ax4.legend(loc='upper right', fontsize=11)
        ax4.grid(True, alpha=0.3, axis='y')
        ax4.set_ylim(0, 1.0)  # 固定y轴范围与Benign一致

        # 添加数值标签和星号
        for i, bar in enumerate(bars_cancer):
            height = bar.get_height()
            p_val = result_cancer['p_values'][i]

            if p_val < 0.001:
                sig = "***"
            elif p_val < 0.01:
                sig = "**"
            elif p_val < 0.05:
                sig = "*"
            else:
                sig = ""

            ax4.text(bar.get_x() + bar.get_width()/2., height,
                    f'{height:.2f}{sig}',
                    ha='center', va='bottom', fontsize=11, fontweight='bold')

        plt.suptitle('CCA Comparison: Benign vs Cancer', fontsize=18, fontweight='bold', y=0.98)

        # 保存图形
        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        filename = f'CA_Comparison_4Panel_BenignVsCancer_{timestamp}'

        for fmt in CCAConfig.SAVE_FORMATS:
            filepath = os.path.join(output_dir, f'{filename}.{fmt}')
            plt.savefig(filepath, dpi=CCAConfig.FIGURE_DPI, bbox_inches='tight', facecolor='white')
            print(f"  [OK] 已保存: {os.path.basename(filepath)}")

        plt.close()

    def plot_top_loadings_comparison(self, output_dir, top_n=5):
        """
        绘制CV1 Top Loadings对照条形图

        2x2布局:
        - (A) Benign: 舌象特征Top Loadings
        - (B) Benign: 脉象特征Top Loadings
        - (C) Cancer: 舌象特征Top Loadings
        - (D) Cancer: 脉象特征Top Loadings

        Parameters:
        -----------
        output_dir : str
            输出目录
        top_n : int, default=5
            显示Top N特征
        """
        print(f"\n[CV1 Top Loadings] 正在生成对照条形图 (Top {top_n})...")

        # 检查是否有两组结果
        if not self.results.get('benign') or not self.results.get('cancer'):
            print("  [SKIP] 缺少必要的分析结果")
            return

        result_benign = self.results['benign']
        result_cancer = self.results['cancer']

        # 创建2x2图形
        fig, axes = plt.subplots(2, 2, figsize=(20, 14))
        fig.suptitle('CV1 Top Loadings Comparison (Benign vs Cancer)',
                    fontsize=18, fontweight='bold', y=0.98)

        # 数据准备
        groups = [
            ('Benign', result_benign),
            ('Cancer', result_cancer)
        ]

        modalities = [
            ('Tongue', 'X_cols', 'x_loadings'),
            ('Pulse', 'Y_cols', 'y_loadings')
        ]

        # 遍历4个子图
        plot_idx = 0
        for group_name, result in groups:
            for modality_name, cols_key, loadings_key in modalities:
                ax = axes[plot_idx // 2, plot_idx % 2]

                # 提取CV1的载荷
                feature_names = result[cols_key]
                loadings = result[loadings_key][:, 0]  # CV1

                # 创建DataFrame并按绝对值排序
                loadings_df = pd.DataFrame({
                    'feature': feature_names,
                    'loading': loadings
                })
                loadings_df['abs_loading'] = loadings_df['loading'].abs()
                loadings_df = loadings_df.sort_values('abs_loading', ascending=False)

                # 取Top N
                top_loadings = loadings_df.head(top_n)

                # 绘制水平条形图
                colors = ['#009688' if x > 0 else '#FF4500' for x in top_loadings['loading']]
                bars = ax.barh(range(len(top_loadings)), top_loadings['loading'], color=colors, alpha=0.8)

                # 设置y轴标签
                ax.set_yticks(range(len(top_loadings)))
                ax.set_yticklabels(top_loadings['feature'], fontsize=11)
                ax.invert_yaxis()  # 最大的在最上面

                # 添加数值标签
                for i, (idx, row) in enumerate(top_loadings.iterrows()):
                    ax.text(row['loading'], i, f'{row["loading"]:.2f}',
                           ha='left' if row['loading'] > 0 else 'right',
                           va='center', fontsize=10, fontweight='bold')

                # 设置标题和标签
                panel_label = ['(A)', '(B)', '(C)', '(D)'][plot_idx]
                ax.set_title(f'{panel_label} {group_name}: {modality_name} Features (CV1)',
                            fontsize=14, fontweight='bold')
                ax.set_xlabel('Loading Coefficient', fontweight='bold', fontsize=12)
                ax.axvline(x=0, color='black', linestyle='-', linewidth=0.8)
                ax.grid(True, alpha=0.3, axis='x')

                plot_idx += 1

        plt.tight_layout()

        # 保存图形
        timestamp = datetime.now().strftime(config.TIMESTAMP_FORMAT)
        filename = f'CA_TopLoadings_CV1_Comparison_{timestamp}'

        for fmt in CCAConfig.SAVE_FORMATS:
            filepath = os.path.join(output_dir, f'{filename}.{fmt}')
            plt.savefig(filepath, dpi=CCAConfig.FIGURE_DPI, bbox_inches='tight', facecolor='white')
            print(f"  [OK] 已保存: {os.path.basename(filepath)}")

        plt.close()

    def run_full_analysis(self, output_dir=None):
        """
        运行完整的CCA分析流程

        Parameters:
        -----------
        output_dir : str, optional
            输出目录，默认使用config配置
        """
        if output_dir is None:
            output_dir = os.path.join(config.OUTPUT_BASE_DIR, 'cca_results')

        # 创建输出目录
        for group in ['benign', 'cancer']:
            group_dir = os.path.join(output_dir, group)
            os.makedirs(group_dir, exist_ok=True)

        print("\n" + "="*60)
        print("开始完整CCA分析")
        print("="*60)

        # 分析所有组
        self.analyze_all_groups()

        # 对每个组生成图表和表格
        for group_name in ['benign', 'cancer']:
            if self.results.get(group_name) is None:
                continue

            group_label = "Benign" if group_name == 'benign' else "Cancer"
            print(f"\n{'='*60}")
            print(f"Generating figures and tables for {group_label} group")
            print('='*60)

            group_dir = os.path.join(output_dir, group_name)

            self.plot_scatter_bar(group_name, group_dir)
            self.plot_loadings_heatmap(group_name, group_dir)
            self.export_summary_table(group_name, group_dir)
            self.export_loadings_tables(group_name, group_dir)  # 新增：导出loadings数值表

        # ========== 新增: 生成对照图 ==========
        print("\n" + "="*60)
        print("正在生成对照图...")
        print("="*60)

        # 生成对照四联图
        self.plot_combined_comparison(output_dir)

        # 生成CV1 Top Loadings对照条形图
        self.plot_top_loadings_comparison(output_dir, top_n=5)
        # =======================================

        print("\n" + "="*60)
        print("[OK] CCA分析完成！")
        print("="*60)
        print(f"\n所有结果已保存到: {output_dir}")


# =============================================================================
# 便捷函数
# =============================================================================

def run_cca_analysis(data, target_col='Group', output_dir=None):
    """
    一键运行CCA分析

    Parameters:
    -----------
    data : pd.DataFrame
        完整数据集
    target_col : str, default='Group'
        目标列名
    output_dir : str, optional
        输出目录

    Returns:
    --------
    analyzer : GroupedCCAAnalyzer
        CCA分析器对象（包含所有结果）
    """
    analyzer = GroupedCCAAnalyzer(data, target_col)
    analyzer.run_full_analysis(output_dir)
    return analyzer


# =============================================================================
# 主程序测试
# =============================================================================

if __name__ == "__main__":
    print("\n" + "="*60)
    print("模块2测试：典型相关分析（CCA）")
    print("="*60)

    try:
        # 加载数据
        from modules.data_loader import DataLoader
        loader = DataLoader()
        loader.load_data()
        loader.auto_detect_features()

        # 运行CCA分析
        analyzer = run_cca_analysis(loader.data)

        print("\n[OK] 测试完成")

    except Exception as e:
        print(f"\n[X] 测试失败: {str(e)}")
        import traceback
        traceback.print_exc()
